import os
import json
import time
import secrets
import threading
import io
import tempfile
import hashlib
import hmac
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any, Tuple

from fastapi import FastAPI, Header, HTTPException, Depends, Request, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, Response

from pydantic import BaseModel

from sqlalchemy import (
    create_engine,
    Column,
    String,
    DateTime,
    Float,
    Integer,
    ForeignKey,
    Text,
    Index,
    UniqueConstraint,
    text,
)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, Session


EXPORT_LIBS_OK = True
EXPORT_LIBS_ERR = ""

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib import colors as rl_colors

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception as e:
    EXPORT_LIBS_OK = False
    EXPORT_LIBS_ERR = str(e)

# Zona horaria (Bogotá)
try:
    from zoneinfo import ZoneInfo  # py>=3.9
except Exception:
    ZoneInfo = None

BOGOTA_TZ = os.getenv("APP_TZ", "America/Bogota")

# -----------------------------
# Configuración
# -----------------------------
DB_URL = os.getenv("DB_URL", "sqlite:///./monitor.db")

# Dashboard default admin (seed user)
DASH_USER = os.getenv("DASH_USER", "admin")
DASH_PASS = os.getenv("DASH_PASS", "admin123")  # change in run_server.bat

# Session expiration (minutes)
SESSION_TTL_MIN = int(os.getenv("SESSION_TTL_MIN", "720"))  # 12h

# Cleanup retention
RETENTION_DAYS = int(os.getenv("RETENTION_DAYS", "7"))
CLEANUP_EVERY_SECONDS = int(os.getenv("CLEANUP_EVERY_SECONDS", str(6 * 60 * 60)))

# Online/offline determination
OFFLINE_AFTER_SECONDS_DEFAULT = int(os.getenv("OFFLINE_AFTER_SECONDS", "60"))

# Export defaults
EXPORT_MAX_POINTS = int(os.getenv("EXPORT_MAX_POINTS", "2000"))  # safety

# Password hash settings
PWD_ITER = int(os.getenv("PWD_ITER", "150000"))

# Static folder
STATIC_DIR = os.getenv("STATIC_DIR", "static")

# -----------------------------
# Email / SMTP
# -----------------------------
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USER or "")
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "1") == "1"

VERIFY_CODE_TTL_MIN = int(os.getenv("VERIFY_CODE_TTL_MIN", "10"))
RESET_CODE_TTL_MIN = int(os.getenv("RESET_CODE_TTL_MIN", "10"))

# -----------------------------
# DB setup
# -----------------------------
engine = create_engine(DB_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()


def now_utc_aware() -> datetime:
    return datetime.now(timezone.utc)


def bogota_tz():
    if ZoneInfo is None:
        return None
    try:
        return ZoneInfo(BOGOTA_TZ)
    except Exception:
        return None


def to_bogota(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    tz = bogota_tz()
    if tz is None:
        return dt.astimezone(timezone(timedelta(hours=-5)))
    return dt.astimezone(tz)


def bogota_to_utc(dt_local: datetime) -> datetime:
    """Interpret dt_local as Bogotá time (if naive) and convert to UTC aware."""
    tz = bogota_tz()
    if dt_local.tzinfo is None:
        if tz is None:
            dt_local = dt_local.replace(tzinfo=timezone(timedelta(hours=-5)))
        else:
            dt_local = dt_local.replace(tzinfo=tz)
    return dt_local.astimezone(timezone.utc)


# -----------------------------
# Password hashing (PBKDF2) - no external deps
# -----------------------------
def _pbkdf2_hash(password: str, salt: bytes, iterations: int) -> bytes:
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations, dklen=32)


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    dk = _pbkdf2_hash(password, salt, PWD_ITER)
    # store as iter$salthex$hashhex
    return f"{PWD_ITER}${salt.hex()}${dk.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        parts = stored.split("$")
        it = int(parts[0])
        salt = bytes.fromhex(parts[1])
        expected = bytes.fromhex(parts[2])
        got = _pbkdf2_hash(password, salt, it)
        return hmac.compare_digest(got, expected)
    except Exception:
        return False


def generate_6digit_code() -> str:
    return f"{secrets.randbelow(1000000):06d}"


def send_email_code(to_email: str, subject: str, title: str, code: str, ttl_min: int):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS or not SMTP_FROM:
        raise RuntimeError(
            "SMTP no configurado. Define SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS y SMTP_FROM."
        )

    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background:#0b0f17; color:#eaf2ff; padding:24px;">
        <div style="max-width:560px; margin:auto; background:#111827; border:1px solid #1f2a3a; border-radius:16px; padding:24px;">
          <h2 style="margin-top:0;">{title}</h2>
          <p>Tu código es:</p>
          <div style="font-size:32px; font-weight:bold; letter-spacing:6px; margin:20px 0; color:#7dd3fc;">
            {code}
          </div>
          <p>Este código vence en {ttl_min} minutos.</p>
          <p>Si tú no hiciste esta solicitud, puedes ignorar este correo.</p>
          <hr style="border:none; border-top:1px solid #1f2a3a; margin:24px 0;">
          <small>SysPulse</small>
        </div>
      </body>
    </html>
    """

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = to_email
    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
        if SMTP_USE_TLS:
            server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_FROM, [to_email], msg.as_string())


def is_code_valid(stored_code: Optional[str], expires_at: Optional[datetime], incoming_code: str) -> bool:
    if not stored_code or not expires_at or not incoming_code:
        return False
    exp = expires_at if expires_at.tzinfo else expires_at.replace(tzinfo=timezone.utc)
    if now_utc_aware() > exp:
        return False
    return secrets.compare_digest((stored_code or "").strip(), (incoming_code or "").strip())


# -----------------------------
# Models
# -----------------------------
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, autoincrement=True)
    username = Column(String, nullable=False, unique=True, index=True)
    email = Column(String, nullable=True, unique=True, index=True)
    password_hash = Column(String, nullable=False)
    role = Column(String, default="user")  # "admin" or "user"

    email_verified = Column(Integer, default=0)  # 0 = no, 1 = sí
    verify_code = Column(String, nullable=True)
    verify_code_expires_at = Column(DateTime(timezone=True), nullable=True)

    reset_code = Column(String, nullable=True)
    reset_code_expires_at = Column(DateTime(timezone=True), nullable=True)

    created_at = Column(DateTime(timezone=True), default=now_utc_aware)

    __table_args__ = (
        UniqueConstraint("username", name="uq_users_username"),
        UniqueConstraint("email", name="uq_users_email"),
    )


class Device(Base):
    __tablename__ = "devices"
    device_id = Column(String, primary_key=True)
    hostname = Column(String, index=True)
    os = Column(String)
    token = Column(String)
    last_seen = Column(DateTime(timezone=True), index=True)

    # ✅ NUEVO
    location = Column(String, default="", nullable=False)

    # ✅ NUEVO MAPA: coordenadas manuales del equipo
    latitude = Column(Float, nullable=True)
    longitude = Column(Float, nullable=True)

    metrics = relationship("Metric", back_populates="device", cascade="all, delete-orphan")


class Metric(Base):
    __tablename__ = "metrics"
    id = Column(Integer, primary_key=True, autoincrement=True)

    device_id = Column(String, ForeignKey("devices.device_id"), index=True)
    ts = Column(DateTime(timezone=True), index=True)

    cpu = Column(Float)
    ram_used = Column(Float)
    ram_total = Column(Float)
    disk_used = Column(Float)
    disk_total = Column(Float)
    uptime_sec = Column(Float)

    # JSON string: {"top_cpu":[{...}], "top_ram":[{...}]}
    processes_json = Column(Text)

    device = relationship("Device", back_populates="metrics")


Index("ix_metrics_device_ts", Metric.device_id, Metric.ts)

Base.metadata.create_all(bind=engine)


def ensure_user_auth_columns():
    db = SessionLocal()
    try:
        rows = db.execute(text("PRAGMA table_info(users)")).fetchall()
        existing = {r[1] for r in rows}

        alters = []
        if "email" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN email VARCHAR")
        if "email_verified" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN email_verified INTEGER DEFAULT 0")
        if "verify_code" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN verify_code VARCHAR")
        if "verify_code_expires_at" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN verify_code_expires_at DATETIME")
        if "reset_code" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN reset_code VARCHAR")
        if "reset_code_expires_at" not in existing:
            alters.append("ALTER TABLE users ADD COLUMN reset_code_expires_at DATETIME")

        for sql in alters:
            db.execute(text(sql))

        db.commit()
    except Exception as e:
        print("[auth] ensure_user_auth_columns error:", e)
        db.rollback()
    finally:
        db.close()


ensure_user_auth_columns()


def ensure_device_map_columns():
    """
    Agrega columnas de mapa a la tabla devices si la base SQLite ya existía.
    Esto permite guardar latitud y longitud sin borrar datos anteriores.
    """
    db = SessionLocal()
    try:
        rows = db.execute(text("PRAGMA table_info(devices)")).fetchall()
        existing = {r[1] for r in rows}

        alters = []
        if "latitude" not in existing:
            alters.append("ALTER TABLE devices ADD COLUMN latitude FLOAT")
        if "longitude" not in existing:
            alters.append("ALTER TABLE devices ADD COLUMN longitude FLOAT")

        for sql in alters:
            db.execute(text(sql))

        db.commit()
    except Exception as e:
        print("[devices] ensure_device_map_columns error:", e)
        db.rollback()
    finally:
        db.close()


ensure_device_map_columns()


def get_db() -> Session:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def seed_admin_user():
    """Create admin user if not exists."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.username == DASH_USER).first()
        if u is None:
            u = User(
                username=DASH_USER,
                email=None,
                password_hash=hash_password(DASH_PASS),
                role="admin",
                email_verified=1,
                created_at=now_utc_aware(),
            )
            db.add(u)
            db.commit()
            print(f"[auth] admin user created: {DASH_USER}")
        else:
            print(f"[auth] admin user exists: {DASH_USER}")
    except Exception as e:
        print("[auth] seed error:", e)
    finally:
        db.close()


seed_admin_user()

# -----------------------------
# FastAPI app
# -----------------------------
app = FastAPI(title="PC Monitor Server")

# Static dashboard
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def _static_path(*parts: str) -> str:
    return os.path.join(STATIC_DIR, *parts)


def _first_existing(paths: List[str]) -> Optional[str]:
    for p in paths:
        if p and os.path.exists(p):
            return p
    return None


@app.get("/")
def root():
    p = _first_existing([_static_path("app.html"), _static_path("app.htm")])
    if p:
        return FileResponse(
            p,
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            }
        )
    return {"status": "ok", "hint": f"Create {STATIC_DIR}/app.html for the dashboard UI"}


@app.get("/favicon.ico")
def favicon_root():
    p = _first_existing([
        _static_path("Favicon.ico"),
        _static_path("favicon.ico"),
        _static_path("Favicon.png"),
        _static_path("favicon.png"),
    ])
    if not p:
        return Response(status_code=404)
    return FileResponse(p)


@app.get("/Favicon.ico")
def favicon_caps():
    p = _first_existing([_static_path("Favicon.ico"), _static_path("favicon.ico")])
    if not p:
        return Response(status_code=404)
    return FileResponse(p)


# -----------------------------
# Error helper (better debugging)
# -----------------------------
@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    return JSONResponse(
        status_code=500,
        content={
            "detail": "Internal Server Error",
            "error": str(exc),
            "path": str(request.url),
        },
    )


# -----------------------------
# Dashboard auth (Bearer session tokens)
# -----------------------------
bearer = HTTPBearer(auto_error=False)

# token -> {"user":..., "created":..., "exp":..., "role":...}
SESSION_TOKENS: Dict[str, Dict[str, Any]] = {}


def _purge_expired_sessions():
    now = now_utc_aware()
    dead = []
    for t, v in SESSION_TOKENS.items():
        try:
            exp = datetime.fromisoformat(v.get("exp"))
            if exp.tzinfo is None:
                exp = exp.replace(tzinfo=timezone.utc)
            if now >= exp:
                dead.append(t)
        except Exception:
            dead.append(t)
    for t in dead:
        SESSION_TOKENS.pop(t, None)


def require_dashboard_auth(
    creds: HTTPAuthorizationCredentials = Depends(bearer),
) -> Dict[str, Any]:
    _purge_expired_sessions()
    if creds is None or creds.scheme.lower() != "bearer":
        raise HTTPException(status_code=401, detail="Missing auth")
    token = creds.credentials.strip()
    if token not in SESSION_TOKENS:
        raise HTTPException(status_code=403, detail="Invalid session")
    return SESSION_TOKENS[token]


def require_admin(sess: Dict[str, Any] = Depends(require_dashboard_auth)) -> Dict[str, Any]:
    if sess.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin required")
    return sess


class LoginIn(BaseModel):
    username: str
    password: str


@app.post("/api/login")
def login(payload: LoginIn, db: Session = Depends(get_db)):
    login_value = (payload.username or "").strip()
    password = payload.password or ""

    if not login_value or not password:
        raise HTTPException(status_code=400, detail="Missing username/email or password")

    u = (
        db.query(User)
        .filter((User.username == login_value) | (User.email == login_value))
        .first()
    )

    if u is None or not verify_password(password, u.password_hash):
        raise HTTPException(status_code=401, detail="Bad credentials")

    if (u.role or "user") != "admin" and int(u.email_verified or 0) != 1:
        raise HTTPException(status_code=403, detail="Debes verificar tu correo antes de iniciar sesión")

    token = secrets.token_urlsafe(32)
    created = now_utc_aware()
    exp = created + timedelta(minutes=SESSION_TTL_MIN)
    SESSION_TOKENS[token] = {
        "user": u.username,
        "email": u.email,
        "role": u.role or "user",
        "created": created.isoformat(),
        "exp": exp.isoformat(),
    }
    return {
        "token": token,
        "user": u.username,
        "email": u.email,
        "role": u.role or "user",
        "exp": exp.isoformat(),
    }


@app.post("/api/logout")
def logout(_sess: Dict[str, Any] = Depends(require_dashboard_auth), creds: HTTPAuthorizationCredentials = Depends(bearer)):
    if creds and creds.credentials in SESSION_TOKENS:
        SESSION_TOKENS.pop(creds.credentials, None)
    return {"status": "ok"}


@app.get("/api/whoami")
def whoami(sess: Dict[str, Any] = Depends(require_dashboard_auth)):
    return {
        "user": sess.get("user", "admin"),
        "email": sess.get("email"),
        "role": sess.get("role", "user"),
        "created": sess.get("created"),
        "exp": sess.get("exp"),
    }


@app.get("/api/health")
def health(db: Session = Depends(get_db)):
    _ = db.query(Device).count()
    return {
        "status": "ok",
        "tz": BOGOTA_TZ,
        "retention_days": RETENTION_DAYS,
        "offline_after_seconds": OFFLINE_AFTER_SECONDS_DEFAULT,
        "export_libs_ok": EXPORT_LIBS_OK,
        "export_libs_error": ("" if EXPORT_LIBS_OK else EXPORT_LIBS_ERR),
        "static_dir": STATIC_DIR,
    }


class CreateUserIn(BaseModel):
    username: str
    password: str
    role: Optional[str] = "user"


class RegisterIn(BaseModel):
    username: str
    email: str
    password: str


class VerifyEmailIn(BaseModel):
    email: str
    code: str


class ForgotPasswordIn(BaseModel):
    email: str


class ResetPasswordIn(BaseModel):
    email: str
    code: str
    new_password: str


@app.post("/api/users", dependencies=[Depends(require_admin)])
def create_user(payload: CreateUserIn, db: Session = Depends(get_db)):
    username = (payload.username or "").strip()
    password = payload.password or ""
    role = (payload.role or "user").strip().lower()
    if role not in ("admin", "user"):
        role = "user"
    if not username or not password:
        raise HTTPException(status_code=400, detail="username/password required")
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=409, detail="User already exists")
    u = User(username=username, password_hash=hash_password(password), role=role, created_at=now_utc_aware())
    db.add(u)
    db.commit()
    return {"status": "ok", "username": username, "role": role}


@app.post("/api/register")
def register(payload: RegisterIn, db: Session = Depends(get_db)):
    username = (payload.username or "").strip()
    email = (payload.email or "").strip().lower()
    password = payload.password or ""

    if not username or not email or not password:
        raise HTTPException(status_code=400, detail="username, email and password are required")

    if "@" not in email:
        raise HTTPException(status_code=400, detail="Correo inválido")

    if len(password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")

    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=409, detail="El nombre de usuario ya existe")

    if db.query(User).filter(User.email == email).first():
        raise HTTPException(status_code=409, detail="El correo ya está registrado")

    code = generate_6digit_code()
    expires = now_utc_aware() + timedelta(minutes=VERIFY_CODE_TTL_MIN)

    u = User(
        username=username,
        email=email,
        password_hash=hash_password(password),
        role="user",
        email_verified=0,
        verify_code=code,
        verify_code_expires_at=expires,
        created_at=now_utc_aware(),
    )
    db.add(u)
    db.commit()

    try:
        send_email_code(
            to_email=email,
            subject="SysPulse - Código de verificación",
            title="Verifica tu cuenta",
            code=code,
            ttl_min=VERIFY_CODE_TTL_MIN,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo enviar el correo: {str(e)}")

    return {
        "status": "ok",
        "message": "Usuario registrado. Revisa tu correo para verificar la cuenta.",
        "email": email,
    }


@app.post("/api/verify-email")
def verify_email(payload: VerifyEmailIn, db: Session = Depends(get_db)):
    email = (payload.email or "").strip().lower()
    code = (payload.code or "").strip()

    if not email or not code:
        raise HTTPException(status_code=400, detail="email and code are required")

    u = db.query(User).filter(User.email == email).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if int(u.email_verified or 0) == 1:
        return {"status": "ok", "message": "El correo ya estaba verificado"}

    if not is_code_valid(u.verify_code, u.verify_code_expires_at, code):
        raise HTTPException(status_code=400, detail="Código inválido o vencido")

    u.email_verified = 1
    u.verify_code = None
    u.verify_code_expires_at = None
    db.commit()

    return {"status": "ok", "message": "Correo verificado correctamente"}


@app.post("/api/resend-verification")
def resend_verification(payload: ForgotPasswordIn, db: Session = Depends(get_db)):
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")

    u = db.query(User).filter(User.email == email).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if int(u.email_verified or 0) == 1:
        return {"status": "ok", "message": "El correo ya está verificado"}

    code = generate_6digit_code()
    expires = now_utc_aware() + timedelta(minutes=VERIFY_CODE_TTL_MIN)

    u.verify_code = code
    u.verify_code_expires_at = expires
    db.commit()

    try:
        send_email_code(
            to_email=email,
            subject="SysPulse - Nuevo código de verificación",
            title="Nuevo código de verificación",
            code=code,
            ttl_min=VERIFY_CODE_TTL_MIN,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo enviar el correo: {str(e)}")

    return {"status": "ok", "message": "Se envió un nuevo código de verificación"}


@app.post("/api/forgot-password")
def forgot_password(payload: ForgotPasswordIn, db: Session = Depends(get_db)):
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")

    u = db.query(User).filter(User.email == email).first()
    if not u:
        return {"status": "ok", "message": "Si el correo existe, se enviará un código de recuperación"}

    code = generate_6digit_code()
    expires = now_utc_aware() + timedelta(minutes=RESET_CODE_TTL_MIN)

    u.reset_code = code
    u.reset_code_expires_at = expires
    db.commit()

    try:
        send_email_code(
            to_email=email,
            subject="SysPulse - Recuperación de contraseña",
            title="Recupera tu contraseña",
            code=code,
            ttl_min=RESET_CODE_TTL_MIN,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo enviar el correo: {str(e)}")

    return {"status": "ok", "message": "Si el correo existe, se enviará un código de recuperación"}


@app.post("/api/reset-password")
def reset_password(payload: ResetPasswordIn, db: Session = Depends(get_db)):
    email = (payload.email or "").strip().lower()
    code = (payload.code or "").strip()
    new_password = payload.new_password or ""

    if not email or not code or not new_password:
        raise HTTPException(status_code=400, detail="email, code and new_password are required")

    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")

    u = db.query(User).filter(User.email == email).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if not is_code_valid(u.reset_code, u.reset_code_expires_at, code):
        raise HTTPException(status_code=400, detail="Código inválido o vencido")

    u.password_hash = hash_password(new_password)
    u.reset_code = None
    u.reset_code_expires_at = None
    db.commit()

    return {"status": "ok", "message": "Contraseña actualizada correctamente"}


# -----------------------------
# Schemas
# -----------------------------
class ProcItem(BaseModel):
    pid: int
    name: str
    cpu: float
    rss: float  # bytes


class ReportPayload(BaseModel):
    device_id: str
    hostname: str
    os: str
    cpu: float
    ram_used: float
    ram_total: float
    disk_used: float
    disk_total: float
    uptime_sec: float
    ip: Optional[str] = None

    # ✅ Opcional: si en el futuro el agente envía ubicación, el server puede recibirla
    latitude: Optional[float] = None
    longitude: Optional[float] = None

    processes_top_cpu: Optional[List[ProcItem]] = None
    processes_top_ram: Optional[List[ProcItem]] = None


class DeviceOut(BaseModel):
    device_id: str
    hostname: str
    os: str
    last_seen: Optional[datetime]
    last_seen_bogota: Optional[datetime]
    status: str
    location: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None


class LocationIn(BaseModel):
    location: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None


class MetricOut(BaseModel):
    ts: datetime
    ts_bogota: datetime
    cpu: float
    ram_used: float
    ram_total: float
    disk_used: float
    disk_total: float
    uptime_sec: float
    processes_json: Optional[str] = None


class LatestOut(BaseModel):
    device_id: str
    ts: datetime
    ts_bogota: datetime
    cpu: float
    ram_used: float
    ram_total: float
    disk_used: float
    disk_total: float
    uptime_sec: float
    processes_top_cpu: List[dict]
    processes_top_ram: List[dict]


class DashboardDeviceSummary(BaseModel):
    device_id: str
    hostname: str
    os: Optional[str] = ""
    status: str
    location: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    last_seen: Optional[datetime] = None
    last_seen_bogota: Optional[datetime] = None
    has_metrics: bool = False
    cpu: float = 0.0
    ram_pct: float = 0.0
    disk_pct: float = 0.0
    uptime_sec: float = 0.0
    risk: float = 0.0


class DashboardSummaryOut(BaseModel):
    status: str
    generated_at: datetime
    generated_at_bogota: datetime
    total_devices: int
    online: int
    offline: int
    with_metrics: int
    with_coordinates: int
    avg_cpu_online: float
    avg_ram_online: float
    avg_disk_online: float
    avg_cpu_all: float
    avg_ram_all: float
    avg_disk_all: float
    health_score: float
    health_label: str
    health_description: str
    top_risk: List[DashboardDeviceSummary]
    top_cpu: List[DashboardDeviceSummary]
    top_ram: List[DashboardDeviceSummary]
    top_disk: List[DashboardDeviceSummary]
    devices: List[DashboardDeviceSummary]


# -----------------------------
# ✅ Fix “solo veo 2 PCs”
# -----------------------------
def _device_id_collision_guard(dev: Device, payload: ReportPayload) -> None:
    """
    Problema típico:
      Copiaste C:\\pc-agent a varios PCs con el mismo device_id.txt,
      entonces todos reportan con el MISMO device_id y en la UI parece que solo hay 1-2 PCs.

    Solución:
      - Borrar device_id.txt en cada PC clonada para que se genere uno nuevo.
      - O reinstalar agente por PC.

    Guard del server:
      Si ya existe ese device_id, y el hostname cambió,
      y el último reporte fue reciente, devolvemos 409 con mensaje claro.
    """
    if not dev:
        return
    existing_host = (dev.hostname or "").strip().lower()
    incoming_host = (payload.hostname or "").strip().lower()
    if not existing_host or not incoming_host:
        return
    if existing_host == incoming_host:
        return

    if dev.last_seen:
        base = dev.last_seen if dev.last_seen.tzinfo else dev.last_seen.replace(tzinfo=timezone.utc)
        delta = (now_utc_aware() - base).total_seconds()
        if delta <= (OFFLINE_AFTER_SECONDS_DEFAULT * 3):
            raise HTTPException(
                status_code=409,
                detail=(
                    "COLISIÓN de device_id detectada. "
                    f"El device_id '{dev.device_id}' ya está asociado a hostname '{dev.hostname}', "
                    f"pero ahora reporta como '{payload.hostname}'.\n\n"
                    "Esto ocurre cuando copiaste la carpeta del agente a varios PCs con el mismo device_id.txt.\n"
                    "✅ Arreglo: en el PC que está mal, borra 'device_id.txt' y reinicia el agente "
                    "(se generará un UUID nuevo)."
                ),
            )


# -----------------------------
# Agent endpoint (NO dashboard auth)
# -----------------------------
@app.post("/api/report")
def report(payload: ReportPayload, x_agent_token: str = Header(default=""), db: Session = Depends(get_db)):
    if not x_agent_token:
        raise HTTPException(status_code=401, detail="Missing X-Agent-Token")

    dev = db.query(Device).filter(Device.device_id == payload.device_id).first()

    if dev is not None:
        _device_id_collision_guard(dev, payload)

    if dev is None:
        dev = Device(
            device_id=payload.device_id,
            hostname=payload.hostname,
            os=payload.os,
            token=x_agent_token,
            last_seen=now_utc_aware(),
            latitude=payload.latitude,
            longitude=payload.longitude,
        )
        db.add(dev)
        db.commit()
        db.refresh(dev)
    else:
        if dev.token != x_agent_token:
            raise HTTPException(status_code=403, detail="Invalid token")
        dev.hostname = payload.hostname
        dev.os = payload.os
        dev.last_seen = now_utc_aware()

        # ✅ Si el agente envía coordenadas, actualiza el mapa automáticamente
        if payload.latitude is not None:
            dev.latitude = payload.latitude
        if payload.longitude is not None:
            dev.longitude = payload.longitude

        db.commit()

    processes_payload = {
        "top_cpu": [p.model_dump() for p in (payload.processes_top_cpu or [])],
        "top_ram": [p.model_dump() for p in (payload.processes_top_ram or [])],
    }

    m = Metric(
        device_id=payload.device_id,
        ts=now_utc_aware(),
        cpu=float(payload.cpu),
        ram_used=float(payload.ram_used),
        ram_total=float(payload.ram_total),
        disk_used=float(payload.disk_used),
        disk_total=float(payload.disk_total),
        uptime_sec=float(payload.uptime_sec),
        processes_json=json.dumps(processes_payload),
    )
    db.add(m)
    db.commit()

    return {"status": "ok", "device_id": dev.device_id, "last_seen_utc": dev.last_seen.isoformat()}


# -----------------------------
# Dashboard summary helpers
# -----------------------------
def _device_status(dev: Device, now: Optional[datetime] = None) -> str:
    now = now or now_utc_aware()
    if dev.last_seen is None:
        return "offline"
    base = dev.last_seen if dev.last_seen.tzinfo else dev.last_seen.replace(tzinfo=timezone.utc)
    delta = (now - base).total_seconds()
    return "online" if delta <= OFFLINE_AFTER_SECONDS_DEFAULT else "offline"


def _avg(values: List[float]) -> float:
    values = [float(v or 0) for v in values]
    return sum(values) / len(values) if values else 0.0


def _latest_metric_for_device(db: Session, device_id: str) -> Optional[Metric]:
    return (
        db.query(Metric)
        .filter(Metric.device_id == device_id)
        .order_by(Metric.ts.desc())
        .first()
    )


def _risk_score(status: str, cpu_v: float, ram_v: float, disk_v: float) -> float:
    status_penalty = 0.0 if status == "online" else 25.0
    risk = (float(cpu_v or 0) * 0.30) + (float(ram_v or 0) * 0.35) + (float(disk_v or 0) * 0.20) + status_penalty
    return max(0.0, min(100.0, risk))


def _health_from_summary(online_count: int, offline_count: int, cpu_online: float, ram_online: float, disk_online: float) -> Tuple[float, str, str]:
    total = online_count + offline_count
    offline_penalty = (offline_count / total) * 35.0 if total else 0.0
    usage_penalty = (cpu_online * 0.20) + (ram_online * 0.25) + (disk_online * 0.15)
    score = max(0.0, min(100.0, 100.0 - offline_penalty - usage_penalty))

    if score < 50:
        return score, "Crítico", "Se recomienda revisar equipos offline y recursos saturados."
    if score < 75:
        return score, "Atención", "Hay consumo alto o equipos offline que conviene revisar."
    return score, "Estable", "El parque de equipos se encuentra en buen estado general."


def _build_dashboard_device_summary(dev: Device, latest_metric: Optional[Metric], now: datetime) -> DashboardDeviceSummary:
    status = _device_status(dev, now)
    has_metrics = latest_metric is not None

    cpu_v = float(latest_metric.cpu or 0) if latest_metric else 0.0
    ram_v = pct(latest_metric.ram_used or 0, latest_metric.ram_total or 0) if latest_metric else 0.0
    disk_v = pct(latest_metric.disk_used or 0, latest_metric.disk_total or 0) if latest_metric else 0.0
    uptime_v = float(latest_metric.uptime_sec or 0) if latest_metric else 0.0
    risk_v = _risk_score(status, cpu_v, ram_v, disk_v) if has_metrics else (25.0 if status == "offline" else 0.0)

    return DashboardDeviceSummary(
        device_id=dev.device_id,
        hostname=dev.hostname or dev.device_id,
        os=dev.os or "",
        status=status,
        location=dev.location or "",
        latitude=dev.latitude,
        longitude=dev.longitude,
        last_seen=dev.last_seen,
        last_seen_bogota=to_bogota(dev.last_seen),
        has_metrics=has_metrics,
        cpu=cpu_v,
        ram_pct=ram_v,
        disk_pct=disk_v,
        uptime_sec=uptime_v,
        risk=risk_v,
    )


# Dashboard
@app.get("/api/devices", response_model=List[DeviceOut])
def list_devices(_sess: Dict[str, Any] = Depends(require_dashboard_auth), db: Session = Depends(get_db)):
    rows = db.query(Device).order_by(Device.last_seen.desc()).all()

    now = now_utc_aware()
    out: List[DeviceOut] = []
    for r in rows:
        if r.last_seen is None:
            status = "offline"
        else:
            base = r.last_seen if r.last_seen.tzinfo else r.last_seen.replace(tzinfo=timezone.utc)
            delta = (now - base).total_seconds()
            status = "online" if delta <= OFFLINE_AFTER_SECONDS_DEFAULT else "offline"

        out.append(
            DeviceOut(
                device_id=r.device_id,
                hostname=r.hostname,
                os=r.os,
                last_seen=r.last_seen,
                last_seen_bogota=to_bogota(r.last_seen),
                status=status,
                location=(r.location or ""),
                latitude=r.latitude,
                longitude=r.longitude,
            )
        )
    return out


@app.patch("/api/devices/{device_id}/location")
@app.put("/api/devices/{device_id}/location")
@app.post("/api/devices/{device_id}/location")
def update_location(
    device_id: str,
    body: LocationIn,
    _sess: Dict[str, Any] = Depends(require_dashboard_auth),
    db: Session = Depends(get_db),
):
    dev = db.query(Device).filter(Device.device_id == device_id).first()
    if not dev:
        raise HTTPException(status_code=404, detail="Device not found")

    dev.location = (body.location or "").strip()
    dev.latitude = body.latitude
    dev.longitude = body.longitude

    db.commit()
    db.refresh(dev)

    return {
        "status": "ok",
        "device_id": dev.device_id,
        "location": dev.location,
        "latitude": dev.latitude,
        "longitude": dev.longitude,
    }


@app.get("/api/devices/{device_id}/latest", response_model=LatestOut)
def latest(device_id: str, _sess: Dict[str, Any] = Depends(require_dashboard_auth), db: Session = Depends(get_db)):
    m = (
        db.query(Metric)
        .filter(Metric.device_id == device_id)
        .order_by(Metric.ts.desc())
        .first()
    )
    if not m:
        raise HTTPException(status_code=404, detail="No metrics for device")

    proc = {"top_cpu": [], "top_ram": []}
    try:
        if m.processes_json:
            proc = json.loads(m.processes_json)
    except Exception:
        proc = {"top_cpu": [], "top_ram": []}

    return LatestOut(
        device_id=device_id,
        ts=m.ts,
        ts_bogota=to_bogota(m.ts) or m.ts,
        cpu=m.cpu,
        ram_used=m.ram_used,
        ram_total=m.ram_total,
        disk_used=m.disk_used,
        disk_total=m.disk_total,
        uptime_sec=m.uptime_sec,
        processes_top_cpu=proc.get("top_cpu", []) or [],
        processes_top_ram=proc.get("top_ram", []) or [],
    )


@app.get("/api/devices/{device_id}/metrics", response_model=List[MetricOut])
def device_metrics(
    device_id: str,
    limit: int = 120,
    _sess: Dict[str, Any] = Depends(require_dashboard_auth),
    db: Session = Depends(get_db),
):
    limit = max(10, min(int(limit), 1000))
    rows = (
        db.query(Metric)
        .filter(Metric.device_id == device_id)
        .order_by(Metric.ts.desc())
        .limit(limit)
        .all()
    )
    out = []
    for r in rows:
        out.append(
            MetricOut(
                ts=r.ts,
                ts_bogota=to_bogota(r.ts) or r.ts,
                cpu=r.cpu,
                ram_used=r.ram_used,
                ram_total=r.ram_total,
                disk_used=r.disk_used,
                disk_total=r.disk_total,
                uptime_sec=r.uptime_sec,
                processes_json=r.processes_json,
            )
        )
    return out


@app.delete("/api/devices/{device_id}")
def delete_device(device_id: str, _sess: Dict[str, Any] = Depends(require_dashboard_auth), db: Session = Depends(get_db)):
    dev = db.query(Device).filter(Device.device_id == device_id).first()
    if not dev:
        raise HTTPException(status_code=404, detail="Device not found")

    db.query(Metric).filter(Metric.device_id == device_id).delete()
    db.query(Device).filter(Device.device_id == device_id).delete()
    db.commit()
    return {"status": "deleted", "device_id": device_id}


@app.post("/api/devices/cleanup-duplicates")
def cleanup_duplicates(_sess: Dict[str, Any] = Depends(require_dashboard_auth), db: Session = Depends(get_db)):
    rows = db.query(Device).order_by(Device.hostname.asc(), Device.last_seen.desc()).all()

    keep: Dict[str, str] = {}
    for d in rows:
        if d.hostname and d.hostname not in keep:
            keep[d.hostname] = d.device_id

    keep_ids = set(keep.values())
    to_delete = db.query(Device).filter(Device.device_id.notin_(keep_ids)).all()

    deleted = 0
    for d in to_delete:
        db.query(Metric).filter(Metric.device_id == d.device_id).delete()
        db.query(Device).filter(Device.device_id == d.device_id).delete()
        deleted += 1

    db.commit()
    return {"status": "ok", "deleted_duplicates": deleted, "kept": list(keep_ids)}



@app.get("/api/dashboard/summary", response_model=DashboardSummaryOut)
def dashboard_summary(_sess: Dict[str, Any] = Depends(require_dashboard_auth), db: Session = Depends(get_db)):
    """
    Resumen consolidado para el Dashboard.
    Evita que el navegador tenga que pedir /latest por cada equipo y entrega datos listos para gráficas.
    """
    now = now_utc_aware()
    devices = db.query(Device).order_by(Device.last_seen.desc()).all()

    summaries: List[DashboardDeviceSummary] = []
    for dev in devices:
        latest_metric = _latest_metric_for_device(db, dev.device_id)
        summaries.append(_build_dashboard_device_summary(dev, latest_metric, now))

    online_items = [x for x in summaries if x.status == "online"]
    offline_items = [x for x in summaries if x.status != "online"]
    metric_items = [x for x in summaries if x.has_metrics]
    online_metric_items = [x for x in online_items if x.has_metrics]

    online_count = len(online_items)
    offline_count = len(offline_items)

    avg_cpu_online = _avg([x.cpu for x in online_metric_items])
    avg_ram_online = _avg([x.ram_pct for x in online_metric_items])
    avg_disk_online = _avg([x.disk_pct for x in online_metric_items])

    avg_cpu_all = _avg([x.cpu for x in metric_items])
    avg_ram_all = _avg([x.ram_pct for x in metric_items])
    avg_disk_all = _avg([x.disk_pct for x in metric_items])

    health_score, health_label, health_description = _health_from_summary(
        online_count,
        offline_count,
        avg_cpu_online,
        avg_ram_online,
        avg_disk_online,
    )

    top_risk = sorted(metric_items, key=lambda x: x.risk, reverse=True)[:10]
    top_cpu = sorted(metric_items, key=lambda x: x.cpu, reverse=True)[:10]
    top_ram = sorted(metric_items, key=lambda x: x.ram_pct, reverse=True)[:10]
    top_disk = sorted(metric_items, key=lambda x: x.disk_pct, reverse=True)[:10]

    return DashboardSummaryOut(
        status="ok",
        generated_at=now,
        generated_at_bogota=to_bogota(now) or now,
        total_devices=len(summaries),
        online=online_count,
        offline=offline_count,
        with_metrics=len(metric_items),
        with_coordinates=len([x for x in summaries if x.latitude is not None and x.longitude is not None]),
        avg_cpu_online=avg_cpu_online,
        avg_ram_online=avg_ram_online,
        avg_disk_online=avg_disk_online,
        avg_cpu_all=avg_cpu_all,
        avg_ram_all=avg_ram_all,
        avg_disk_all=avg_disk_all,
        health_score=health_score,
        health_label=health_label,
        health_description=health_description,
        top_risk=top_risk,
        top_cpu=top_cpu,
        top_ram=top_ram,
        top_disk=top_disk,
        devices=summaries,
    )


# ==========================================================
# ✅ EXPORT HELPERS (PDF/XLSX)
# ==========================================================
def ensure_export_libs():
    if not EXPORT_LIBS_OK:
        raise HTTPException(
            status_code=500,
            detail=(
                "Exportación no disponible porque faltan librerías. "
                "Instala dependencias con:\n"
                "  pip install reportlab openpyxl matplotlib tzdata\n\n"
                f"Error detectado: {EXPORT_LIBS_ERR}"
            ),
        )


def parse_date_or_datetime(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    s = s.strip()
    try:
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            y, m, d = map(int, s.split("-"))
            return datetime(y, m, d, 0, 0, 0)
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {s}. Use YYYY-MM-DD or ISO datetime.")


def normalize_range_to_utc(start_s: Optional[str], end_s: Optional[str]) -> Tuple[datetime, datetime]:
    start_dt = parse_date_or_datetime(start_s)
    end_dt = parse_date_or_datetime(end_s)

    if start_dt is None and end_dt is None:
        end_utc = now_utc_aware()
        start_utc = end_utc - timedelta(hours=24)
        return start_utc, end_utc

    if start_dt is not None and end_dt is None:
        day_start_local = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end_local = start_dt.replace(hour=23, minute=59, second=59, microsecond=999000)
        return bogota_to_utc(day_start_local), bogota_to_utc(day_end_local)

    if start_dt is None and end_dt is not None:
        end_local = end_dt
        if len((end_s or "")) == 10:
            end_local = end_dt.replace(hour=23, minute=59, second=59, microsecond=999000)
        end_utc = bogota_to_utc(end_local)
        start_utc = end_utc - timedelta(hours=24)
        return start_utc, end_utc

    assert start_dt is not None and end_dt is not None
    start_local = start_dt
    end_local = end_dt
    if len((end_s or "")) == 10:
        end_local = end_dt.replace(hour=23, minute=59, second=59, microsecond=999000)

    start_utc = bogota_to_utc(start_local)
    end_utc = bogota_to_utc(end_local)

    if end_utc < start_utc:
        raise HTTPException(status_code=400, detail="end must be >= start")

    return start_utc, end_utc


def pct(used: float, total: float) -> float:
    if not total or total <= 0:
        return 0.0
    return float(used) / float(total) * 100.0


def fetch_metrics_range(db: Session, device_id: str, start_utc: datetime, end_utc: datetime) -> List[Metric]:
    rows = (
        db.query(Metric)
        .filter(Metric.device_id == device_id)
        .filter(Metric.ts >= start_utc)
        .filter(Metric.ts <= end_utc)
        .order_by(Metric.ts.asc())
        .limit(EXPORT_MAX_POINTS)
        .all()
    )
    return rows


def render_taskmgr_line_chart(title: str, xs: List[datetime], ys: List[float], y_max: float = 100.0) -> bytes:
    fig_w, fig_h = 10.0, 2.6
    fig = plt.figure(figsize=(fig_w, fig_h), dpi=140)
    ax = fig.add_subplot(111)

    # Reportes en tema claro para PDF/Excel
    bg = "#ffffff"
    grid = "#dbeafe"
    text_color = "#0f172a"

    fig.patch.set_facecolor(bg)
    ax.set_facecolor(bg)

    ax.grid(True, color=grid, linewidth=0.8, alpha=0.8)
    ax.set_axisbelow(True)

    ax.plot(xs, ys, linewidth=2.0)
    ax.fill_between(xs, ys, [0] * len(ys), alpha=0.18)

    ax.set_ylim(0, y_max)
    ax.set_title(title, color=text_color, fontsize=11, fontweight="bold", loc="left")
    ax.tick_params(colors="#475569", labelsize=8)
    for spine in ax.spines.values():
        spine.set_color("#cbd5e1")

    if xs:
        step = max(1, len(xs) // 8)
        xticks = xs[::step]
        ax.set_xticks(xticks)
        labels = []
        for d in xticks:
            try:
                labels.append((to_bogota(d) or d).strftime("%H:%M"))
            except Exception:
                labels.append(d.strftime("%H:%M"))
        ax.set_xticklabels(labels, rotation=0, ha="center")

    ax.set_ylabel("%", color="#475569", fontsize=8)

    buf = io.BytesIO()
    plt.tight_layout()
    fig.savefig(buf, format="png", facecolor=fig.get_facecolor())
    plt.close(fig)
    return buf.getvalue()


def build_summary(rows: List[Metric]) -> Dict[str, float]:
    if not rows:
        return {"cpu_avg": 0, "cpu_max": 0, "ram_avg": 0, "ram_max": 0, "disk_avg": 0, "disk_max": 0}
    cpu_vals = [float(r.cpu or 0) for r in rows]
    ram_vals = [pct(r.ram_used or 0, r.ram_total or 0) for r in rows]
    disk_vals = [pct(r.disk_used or 0, r.disk_total or 0) for r in rows]

    def avg(xs): return sum(xs) / len(xs) if xs else 0.0

    return {
        "cpu_avg": avg(cpu_vals),
        "cpu_max": max(cpu_vals) if cpu_vals else 0.0,
        "ram_avg": avg(ram_vals),
        "ram_max": max(ram_vals) if ram_vals else 0.0,
        "disk_avg": avg(disk_vals),
        "disk_max": max(disk_vals) if disk_vals else 0.0,
    }



def build_report_interpretation(summary: Dict[str, float], rows: List[Metric], device: Optional[Device] = None) -> List[str]:
    """Genera una explicación ejecutiva del comportamiento del equipo en el rango exportado."""
    if not rows:
        return [
            "No se encontraron métricas en el rango seleccionado, por lo tanto no es posible generar una interpretación técnica."
        ]

    cpu_avg = float(summary.get("cpu_avg", 0) or 0)
    cpu_max = float(summary.get("cpu_max", 0) or 0)
    ram_avg = float(summary.get("ram_avg", 0) or 0)
    ram_max = float(summary.get("ram_max", 0) or 0)
    disk_avg = float(summary.get("disk_avg", 0) or 0)
    disk_max = float(summary.get("disk_max", 0) or 0)

    notes: List[str] = []
    host = (device.hostname if device and device.hostname else "el equipo") if device else "el equipo"

    notes.append(
        f"Durante el rango analizado se procesaron {len(rows)} puntos de monitoreo para {host}. "
        f"Los promedios observados fueron CPU {cpu_avg:.1f}%, RAM {ram_avg:.1f}% y Disco {disk_avg:.1f}%."
    )

    if cpu_max >= 90:
        notes.append(
            f"La CPU alcanzó un pico crítico de {cpu_max:.1f}%. Esto puede indicar procesos pesados, tareas en segundo plano o saturación temporal del equipo."
        )
    elif cpu_avg >= 70:
        notes.append(
            f"La CPU mantuvo un promedio alto de {cpu_avg:.1f}%. Se recomienda revisar procesos activos y carga de trabajo."
        )
    else:
        notes.append(
            f"El uso de CPU se mantuvo controlado. El pico máximo fue {cpu_max:.1f}% y el promedio fue {cpu_avg:.1f}%."
        )

    if ram_max >= 90 or ram_avg >= 80:
        notes.append(
            f"La memoria RAM presenta uso elevado: promedio {ram_avg:.1f}% y máximo {ram_max:.1f}%. Conviene validar aplicaciones abiertas, servicios residentes o posible necesidad de ampliar memoria."
        )
    else:
        notes.append(
            f"La memoria RAM se mantiene en un rango aceptable para el periodo revisado, con promedio {ram_avg:.1f}%."
        )

    if disk_max >= 90 or disk_avg >= 85:
        notes.append(
            f"El disco se encuentra cerca de saturación: promedio {disk_avg:.1f}% y máximo {disk_max:.1f}%. Se recomienda liberar espacio, depurar archivos temporales y revisar respaldos."
        )
    elif disk_avg >= 70:
        notes.append(
            f"El disco muestra ocupación considerable ({disk_avg:.1f}% promedio). Aunque no es crítico, debe monitorearse para evitar falta de espacio."
        )
    else:
        notes.append(
            f"La ocupación de disco se encuentra estable. Promedio registrado: {disk_avg:.1f}%."
        )

    if cpu_max >= 90 and ram_max >= 90:
        notes.append(
            "La combinación de picos altos de CPU y RAM puede generar lentitud perceptible para el usuario. Se recomienda revisar el Top de procesos incluido en este informe."
        )

    notes.append(
        "Recomendación general: priorizar revisión si el equipo combina RAM alta, disco alto o picos frecuentes de CPU. "
        "Estos indicadores ayudan a anticipar lentitud, bloqueos o necesidad de mantenimiento preventivo."
    )
    return notes



def latest_processes_from_rows(rows: List[Metric]) -> Tuple[List[dict], List[dict]]:
    if not rows:
        return [], []
    last = rows[-1]
    proc = {"top_cpu": [], "top_ram": []}
    try:
        if last.processes_json:
            proc = json.loads(last.processes_json)
    except Exception:
        proc = {"top_cpu": [], "top_ram": []}
    return proc.get("top_cpu", []) or [], proc.get("top_ram", []) or []


def sanitize_filename(s: str) -> str:
    bad = '<>:"/\\|?*'
    for ch in bad:
        s = s.replace(ch, "_")
    return s.strip()[:120] if s else "device"


def _pick_logo_for_reports() -> Optional[str]:
    return _first_existing([
        _static_path("Logotipo.png"),
        _static_path("logo.png"),
        _static_path("Logotipo.jpg"),
        _static_path("logo.jpg"),
    ])


def make_pdf_report(device: Device, start_utc: datetime, end_utc: datetime, rows: List[Metric]) -> bytes:
    ensure_export_libs()

    hostname = device.hostname or device.device_id
    summary = build_summary(rows)

    xs = [r.ts if (r.ts.tzinfo is not None) else r.ts.replace(tzinfo=timezone.utc) for r in rows]
    cpu = [float(r.cpu or 0) for r in rows]
    ram = [pct(r.ram_used or 0, r.ram_total or 0) for r in rows]
    disk = [pct(r.disk_used or 0, r.disk_total or 0) for r in rows]

    cpu_png = render_taskmgr_line_chart("CPU (%)", xs, cpu)
    ram_png = render_taskmgr_line_chart("RAM (%)", xs, ram)
    disk_png = render_taskmgr_line_chart("DISCO (%)", xs, disk)

    top_cpu, top_ram = latest_processes_from_rows(rows)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # Tema claro para presentación y lectura del informe
    bg = rl_colors.HexColor("#ffffff")
    fg = rl_colors.HexColor("#0f172a")
    muted = rl_colors.HexColor("#475569")
    line = rl_colors.HexColor("#cbd5e1")

    logo_path = _pick_logo_for_reports()

    def header(title: str):
        c.setFillColor(bg)
        c.rect(0, 0, W, H, fill=1, stroke=0)

        # Franja superior visible en tema claro
        c.setFillColor(rl_colors.HexColor("#eaf2ff"))
        c.roundRect(1.6 * cm, H - 2.65 * cm, W - 7.2 * cm, 1.05 * cm, 8, fill=1, stroke=0)

        if logo_path:
            try:
                img = ImageReader(logo_path)
                c.drawImage(img, W - 5.0 * cm, H - 2.6 * cm, width=3.2 * cm, height=3.2 * cm, mask="auto")
            except Exception:
                pass

        c.setFillColor(rl_colors.HexColor("#0f172a"))
        c.setFont("Helvetica-Bold", 16)
        c.drawString(2 * cm, H - 2.0 * cm, title)

        c.setFillColor(rl_colors.HexColor("#1e3a8a"))
        c.setFont("Helvetica", 9)
        c.drawString(2 * cm, H - 2.6 * cm, "SysPulse - Reporte por dispositivo")

        c.setStrokeColor(line)
        c.setLineWidth(1)
        c.line(2 * cm, H - 2.9 * cm, W - 2 * cm, H - 2.9 * cm)

    header(f"Informe: {hostname}")

    c.setFillColor(fg)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, H - 3.6 * cm, "Detalle del equipo")

    c.setFillColor(muted)
    c.setFont("Helvetica", 10)
    c.drawString(2 * cm, H - 4.2 * cm, f"Device ID: {device.device_id}")
    c.drawString(2 * cm, H - 4.7 * cm, f"Hostname: {hostname}")
    c.drawString(2 * cm, H - 5.2 * cm, f"OS: {device.os or '-'}")
    c.drawString(2 * cm, H - 5.7 * cm, f"Ubicación: {device.location or '-'}")
    c.drawString(2 * cm, H - 6.2 * cm, f"Coordenadas: {device.latitude if device.latitude is not None else '-'}, {device.longitude if device.longitude is not None else '-'}")

    s_local = to_bogota(start_utc) or start_utc
    e_local = to_bogota(end_utc) or end_utc
    c.drawString(2 * cm, H - 6.7 * cm, f"Rango: {s_local.strftime('%Y-%m-%d %H:%M')} → {e_local.strftime('%Y-%m-%d %H:%M')} (Bogotá)")
    c.drawString(2 * cm, H - 7.2 * cm, f"Puntos: {len(rows)}")

    y0 = H - 8.2 * cm
    c.setFillColor(fg)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, y0, "Resumen del rango")

    chips = [
        ("CPU prom", f"{summary['cpu_avg']:.1f}%"),
        ("CPU máx", f"{summary['cpu_max']:.1f}%"),
        ("RAM prom", f"{summary['ram_avg']:.1f}%"),
        ("RAM máx", f"{summary['ram_max']:.1f}%"),
        ("DISK prom", f"{summary['disk_avg']:.1f}%"),
        ("DISK máx", f"{summary['disk_max']:.1f}%"),
    ]

    x = 2 * cm
    y = y0 - 0.7 * cm
    for label, val in chips:
        w = 5.2 * cm
        h = 0.85 * cm
        c.setFillColor(rl_colors.HexColor("#f8fafc"))
        c.setStrokeColor(line)
        c.roundRect(x, y - h + 0.1 * cm, w, h, 8, fill=1, stroke=1)
        c.setFillColor(muted)
        c.setFont("Helvetica", 8)
        c.drawString(x + 0.35 * cm, y - 0.45 * cm, label)
        c.setFillColor(fg)
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(x + w - 0.35 * cm, y - 0.45 * cm, val)

        x += w + 0.35 * cm
        if x + w > W - 2 * cm:
            x = 2 * cm
            y -= 1.1 * cm

    def draw_chart(png_bytes: bytes, title: str, y_top: float):
        c.setFillColor(fg)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2 * cm, y_top, title)
        img = ImageReader(io.BytesIO(png_bytes))
        c.drawImage(img, 2 * cm, y_top - 6.2 * cm, width=W - 4 * cm, height=5.6 * cm, preserveAspectRatio=True, mask="auto")

    y_charts = y - 0.8 * cm
    draw_chart(cpu_png, "CPU", y_charts)
    c.showPage()

    header(f"Informe: {hostname} (Gráficas)")
    y_top = H - 3.2 * cm
    draw_chart(ram_png, "RAM", y_top)
    draw_chart(disk_png, "DISCO", y_top - 6.8 * cm)
    c.showPage()

    header(f"Informe: {hostname} (Procesos + Tabla)")
    c.setFillColor(fg)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, H - 3.6 * cm, "Top procesos (último punto del rango)")

    def draw_proc_table(title: str, procs: List[dict], x0: float, y0_: float):
        c.setFillColor(fg)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x0, y0_, title)

        yy = y0_ - 0.6 * cm
        c.setStrokeColor(line)
        c.setFillColor(muted)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x0, yy, "PID")
        c.drawString(x0 + 1.2 * cm, yy, "Proceso")
        c.drawRightString(x0 + 7.8 * cm, yy, "CPU%")
        c.drawRightString(x0 + 10.0 * cm, yy, "RAM")

        yy -= 0.2 * cm
        c.line(x0, yy, x0 + 10.0 * cm, yy)
        yy -= 0.35 * cm

        c.setFont("Helvetica", 8)
        c.setFillColor(rl_colors.HexColor("#0f172a"))

        shown = 0
        for p in (procs or [])[:10]:
            pid = str(p.get("pid", ""))
            name = str(p.get("name", "-"))[:36]
            cpu_v = float(p.get("cpu", 0) or 0)
            rss = float(p.get("rss", 0) or 0)
            ram_gb = rss / (1024 * 1024 * 1024)

            c.drawString(x0, yy, pid)
            c.drawString(x0 + 1.2 * cm, yy, name)
            c.drawRightString(x0 + 7.8 * cm, yy, f"{cpu_v:.1f}")
            c.drawRightString(x0 + 10.0 * cm, yy, f"{ram_gb:.2f} GB")
            yy -= 0.45 * cm
            shown += 1
            if yy < 3.0 * cm:
                break

        if shown == 0:
            c.setFillColor(muted)
            c.drawString(x0, yy, "Sin datos")
            yy -= 0.45 * cm

        return yy

    left_x = 2 * cm
    top_y = H - 4.4 * cm
    y_after_left = draw_proc_table("Top CPU", top_cpu, left_x, top_y)
    y_after_right = draw_proc_table("Top RAM", top_ram, left_x, y_after_left - 0.6 * cm)

    c.setFillColor(fg)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, y_after_right - 0.4 * cm, "Métricas (muestra)")

    sample = rows[-60:] if len(rows) > 60 else rows

    yy = y_after_right - 1.1 * cm
    c.setFillColor(muted)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(2 * cm, yy, "Hora (Bogotá)")
    c.drawRightString(9.2 * cm, yy, "CPU%")
    c.drawRightString(12.0 * cm, yy, "RAM%")
    c.drawRightString(15.0 * cm, yy, "DISK%")
    yy -= 0.25 * cm
    c.setStrokeColor(line)
    c.line(2 * cm, yy, W - 2 * cm, yy)
    yy -= 0.45 * cm

    c.setFillColor(rl_colors.HexColor("#0f172a"))
    c.setFont("Helvetica", 8)
    for r in sample:
        tsb = to_bogota(r.ts) or r.ts
        cpu_v = float(r.cpu or 0)
        ram_v = pct(r.ram_used or 0, r.ram_total or 0)
        disk_v = pct(r.disk_used or 0, r.disk_total or 0)

        c.drawString(2 * cm, yy, tsb.strftime("%Y-%m-%d %H:%M:%S"))
        c.drawRightString(9.2 * cm, yy, f"{cpu_v:.1f}")
        c.drawRightString(12.0 * cm, yy, f"{ram_v:.1f}")
        c.drawRightString(15.0 * cm, yy, f"{disk_v:.1f}")
        yy -= 0.42 * cm
        if yy < 2.2 * cm:
            break

    # Interpretación debajo de Métricas (muestra)
    interpretation = build_report_interpretation(summary, rows, device)
    if yy < 5.2 * cm:
        c.showPage()
        header(f"Informe: {hostname} (Interpretación)")
        yy = H - 3.6 * cm

    c.setFillColor(rl_colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, yy, "Interpretación de resultados")
    yy -= 0.55 * cm

    c.setFillColor(muted)
    c.setFont("Helvetica", 8.6)

    def _wrap_text(txt: str, max_chars: int = 112) -> List[str]:
        words = str(txt).split()
        lines, cur = [], ""
        for w in words:
            if len(cur) + len(w) + 1 <= max_chars:
                cur = (cur + " " + w).strip()
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines

    for idx, item in enumerate(interpretation, start=1):
        lines = _wrap_text(f"{idx}. {item}", 108)
        for line_txt in lines:
            if yy < 2.2 * cm:
                c.showPage()
                header(f"Informe: {hostname} (Interpretación)")
                yy = H - 3.6 * cm
                c.setFillColor(muted)
                c.setFont("Helvetica", 8.6)
            c.drawString(2 * cm, yy, line_txt)
            yy -= 0.38 * cm
        yy -= 0.12 * cm

    c.save()
    return buf.getvalue()


def make_xlsx_report(device: Device, start_utc: datetime, end_utc: datetime, rows: List[Metric]) -> bytes:
    ensure_export_libs()

    hostname = device.hostname or device.device_id
    summary = build_summary(rows)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Resumen"

    title_font = Font(bold=True, size=14)
    h_font = Font(bold=True, size=11)
    muted_fill = PatternFill("solid", fgColor="EAF2FF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws_sum["A1"] = "SysPulse - Informe por PC"
    ws_sum["A1"].font = title_font

    s_local = to_bogota(start_utc) or start_utc
    e_local = to_bogota(end_utc) or end_utc

    ws_sum["A3"] = "Device ID:"
    ws_sum["B3"] = device.device_id
    ws_sum["A4"] = "Hostname:"
    ws_sum["B4"] = hostname
    ws_sum["A5"] = "OS:"
    ws_sum["B5"] = device.os or "-"
    ws_sum["A6"] = "Ubicación:"
    ws_sum["B6"] = device.location or "-"
    ws_sum["A7"] = "Latitud:"
    ws_sum["B7"] = device.latitude
    ws_sum["A8"] = "Longitud:"
    ws_sum["B8"] = device.longitude
    ws_sum["A9"] = "Rango (Bogotá):"
    ws_sum["B9"] = f"{s_local.strftime('%Y-%m-%d %H:%M')} → {e_local.strftime('%Y-%m-%d %H:%M')}"
    ws_sum["A10"] = "Puntos:"
    ws_sum["B10"] = len(rows)

    for r in range(3, 11):
        ws_sum[f"A{r}"].font = h_font
        ws_sum[f"A{r}"].fill = muted_fill
        ws_sum[f"A{r}"].alignment = left
        ws_sum[f"B{r}"].alignment = left

    ws_sum["A12"] = "Resumen"
    ws_sum["A12"].font = Font(bold=True, size=12)

    headers = ["Métrica", "Promedio", "Máximo"]
    ws_sum.append([])
    ws_sum.append(headers)
    for col, val in enumerate(headers, start=1):
        cell = ws_sum.cell(row=14, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    rows_sum = [
        ("CPU (%)", summary["cpu_avg"], summary["cpu_max"]),
        ("RAM (%)", summary["ram_avg"], summary["ram_max"]),
        ("DISCO (%)", summary["disk_avg"], summary["disk_max"]),
    ]
    for i, (name, avg_v, max_v) in enumerate(rows_sum, start=15):
        ws_sum.cell(row=i, column=1, value=name)
        ws_sum.cell(row=i, column=2, value=float(avg_v))
        ws_sum.cell(row=i, column=3, value=float(max_v))

    for col in range(1, 4):
        ws_sum.column_dimensions[get_column_letter(col)].width = 20

    ws = wb.create_sheet("Métricas")
    m_headers = ["TS (Bogotá)", "CPU %", "RAM %", "RAM usado (GB)", "RAM total (GB)", "DISCO %", "DISCO usado (GB)", "DISCO total (GB)", "Uptime (s)"]
    ws.append(m_headers)
    for col, val in enumerate(m_headers, start=1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for r in rows:
        tsb = to_bogota(r.ts) or r.ts
        ram_pct = pct(r.ram_used or 0, r.ram_total or 0)
        disk_pct = pct(r.disk_used or 0, r.disk_total or 0)
        ws.append([
            tsb.strftime("%Y-%m-%d %H:%M:%S"),
            float(r.cpu or 0),
            float(ram_pct),
            float((r.ram_used or 0) / (1024 * 1024 * 1024)),
            float((r.ram_total or 0) / (1024 * 1024 * 1024)),
            float(disk_pct),
            float((r.disk_used or 0) / (1024 * 1024 * 1024)),
            float((r.disk_total or 0) / (1024 * 1024 * 1024)),
            float(r.uptime_sec or 0),
        ])

    for col in range(1, len(m_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Interpretación debajo de la tabla de Métricas
    interpretation = build_report_interpretation(summary, rows, device)
    interp_start = ws.max_row + 3
    ws.cell(row=interp_start, column=1, value="Interpretación de resultados")
    ws.cell(row=interp_start, column=1).font = Font(bold=True, size=12)
    ws.cell(row=interp_start, column=1).fill = PatternFill("solid", fgColor="DBEAFE")
    ws.merge_cells(start_row=interp_start, start_column=1, end_row=interp_start, end_column=9)

    rr = interp_start + 1
    for idx, txt in enumerate(interpretation, start=1):
        ws.cell(row=rr, column=1, value=f"{idx}. {txt}")
        ws.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
        ws.row_dimensions[rr].height = 42
        rr += 1

    ws_interp = wb.create_sheet("Interpretación")
    ws_interp["A1"] = "Interpretación de resultados"
    ws_interp["A1"].font = Font(bold=True, size=14)
    ws_interp["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws_interp["A3"] = "Equipo"
    ws_interp["B3"] = hostname
    ws_interp["A4"] = "Rango"
    ws_interp["B4"] = f"{s_local.strftime('%Y-%m-%d %H:%M')} → {e_local.strftime('%Y-%m-%d %H:%M')}"
    ws_interp["A6"] = "Detalle"
    ws_interp["A6"].font = Font(bold=True)
    for idx, txt in enumerate(interpretation, start=1):
        i = idx + 6
        ws_interp.cell(row=i, column=1, value=f"{idx}. {txt}")
        ws_interp.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws_interp.row_dimensions[i].height = 48
    ws_interp.column_dimensions["A"].width = 120
    ws_interp.column_dimensions["B"].width = 35

    top_cpu, top_ram = latest_processes_from_rows(rows)
    wsp = wb.create_sheet("Procesos")
    wsp["A1"] = "Top Procesos (último punto del rango)"
    wsp["A1"].font = Font(bold=True, size=12)

    def write_proc_block(title: str, procs: List[dict], start_row: int) -> int:
        wsp[f"A{start_row}"] = title
        wsp[f"A{start_row}"].font = Font(bold=True, size=11)
        headers_ = ["PID", "Nombre", "CPU %", "RAM (GB)"]
        for j, h in enumerate(headers_, start=1):
            cell = wsp.cell(row=start_row + 1, column=j, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        row_i = start_row + 2
        if not procs:
            wsp.cell(row=row_i, column=1, value="Sin datos")
            return row_i + 2

        for p in procs[:15]:
            rss = float(p.get("rss", 0) or 0)
            wsp.append([
                int(p.get("pid", 0) or 0),
                str(p.get("name", "-")),
                float(p.get("cpu", 0) or 0),
                float(rss / (1024 * 1024 * 1024)),
            ])
            row_i += 1

        return row_i + 2

    next_row = write_proc_block("Top CPU", top_cpu, 3)
    write_proc_block("Top RAM", top_ram, next_row)

    for col in range(1, 5):
        wsp.column_dimensions[get_column_letter(col)].width = 24

    ws_ch = wb.create_sheet("Gráficas")
    ws_ch["A1"] = "Gráficas (estilo rendimiento)"
    ws_ch["A1"].font = Font(bold=True, size=12)

    xs = [r.ts if (r.ts.tzinfo is not None) else r.ts.replace(tzinfo=timezone.utc) for r in rows]
    cpu_series = [float(r.cpu or 0) for r in rows]
    ram_series = [pct(r.ram_used or 0, r.ram_total or 0) for r in rows]
    disk_series = [pct(r.disk_used or 0, r.disk_total or 0) for r in rows]

    cpu_png = render_taskmgr_line_chart("CPU (%)", xs, cpu_series)
    ram_png = render_taskmgr_line_chart("RAM (%)", xs, ram_series)
    disk_png = render_taskmgr_line_chart("DISCO (%)", xs, disk_series)

    tmp_files = []
    try:
        for idx, (png, anchor) in enumerate([(cpu_png, "A3"), (ram_png, "A20"), (disk_png, "A37")], start=1):
            tf = tempfile.NamedTemporaryFile(delete=False, suffix=f"_chart_{idx}.png")
            tf.write(png)
            tf.flush()
            tf.close()
            tmp_files.append(tf.name)
            img = XLImage(tf.name)
            img.width = 880
            img.height = 240
            ws_ch.add_image(img, anchor)
    finally:
        pass

    out = io.BytesIO()
    wb.save(out)

    for f in tmp_files:
        try:
            os.unlink(f)
        except Exception:
            pass

    return out.getvalue()


# ==========================================================
# ✅ EXPORT ENDPOINTS
# ==========================================================
def _export_normalize_params(
    start: Optional[str],
    end: Optional[str],
    from_: Optional[str],
    to_: Optional[str],
    mode: Optional[str],
) -> Tuple[Optional[str], Optional[str]]:
    if from_ or to_:
        if (mode or "").lower() == "day":
            return from_, None
        return from_, to_
    return start, end


@app.get("/api/devices/{device_id}/export/pdf")
def export_pdf(
    device_id: str,
    start: Optional[str] = Query(default=None, description="YYYY-MM-DD or ISO datetime (Bogotá interpreted)"),
    end: Optional[str] = Query(default=None, description="YYYY-MM-DD or ISO datetime (Bogotá interpreted)"),
    from_: Optional[str] = Query(default=None, alias="from"),
    to_: Optional[str] = Query(default=None, alias="to"),
    mode: Optional[str] = Query(default=None, description="day|range"),
    tz: Optional[str] = Query(default=None, description="timezone name; UI param (informational)"),
    _sess: Dict[str, Any] = Depends(require_dashboard_auth),
    db: Session = Depends(get_db),
):
    dev = db.query(Device).filter(Device.device_id == device_id).first()
    if not dev:
        raise HTTPException(status_code=404, detail="Device not found")

    s, e = _export_normalize_params(start, end, from_, to_, mode)
    start_utc, end_utc = normalize_range_to_utc(s, e)
    rows = fetch_metrics_range(db, device_id, start_utc, end_utc)

    if not rows:
        raise HTTPException(status_code=404, detail="No metrics in selected range")

    pdf_bytes = make_pdf_report(dev, start_utc, end_utc, rows)

    host = sanitize_filename(dev.hostname or dev.device_id)
    s_local = (to_bogota(start_utc) or start_utc).strftime("%Y-%m-%d")
    e_local = (to_bogota(end_utc) or end_utc).strftime("%Y-%m-%d")
    filename = f"SysPulse_{host}_{s_local}_a_{e_local}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/devices/{device_id}/export/xlsx")
def export_xlsx(
    device_id: str,
    start: Optional[str] = Query(default=None, description="YYYY-MM-DD or ISO datetime (Bogotá interpreted)"),
    end: Optional[str] = Query(default=None, description="YYYY-MM-DD or ISO datetime (Bogotá interpreted)"),
    from_: Optional[str] = Query(default=None, alias="from"),
    to_: Optional[str] = Query(default=None, alias="to"),
    mode: Optional[str] = Query(default=None, description="day|range"),
    tz: Optional[str] = Query(default=None, description="timezone name; UI param (informational)"),
    _sess: Dict[str, Any] = Depends(require_dashboard_auth),
    db: Session = Depends(get_db),
):
    dev = db.query(Device).filter(Device.device_id == device_id).first()
    if not dev:
        raise HTTPException(status_code=404, detail="Device not found")

    s, e = _export_normalize_params(start, end, from_, to_, mode)
    start_utc, end_utc = normalize_range_to_utc(s, e)
    rows = fetch_metrics_range(db, device_id, start_utc, end_utc)

    if not rows:
        raise HTTPException(status_code=404, detail="No metrics in selected range")

    xlsx_bytes = make_xlsx_report(dev, start_utc, end_utc, rows)

    host = sanitize_filename(dev.hostname or dev.device_id)
    s_local = (to_bogota(start_utc) or start_utc).strftime("%Y-%m-%d")
    e_local = (to_bogota(end_utc) or end_utc).strftime("%Y-%m-%d")
    filename = f"SysPulse_{host}_{s_local}_a_{e_local}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# -----------------------------
# Cleanup thread: delete old metrics
# -----------------------------
def cleanup_old_metrics_loop():
    while True:
        try:
            db = SessionLocal()
            cutoff = now_utc_aware() - timedelta(days=RETENTION_DAYS)
            deleted = db.query(Metric).filter(Metric.ts < cutoff).delete()
            db.commit()
            db.close()
            print(f"[cleanup] deleted={deleted} metrics older than {RETENTION_DAYS} days")
        except Exception as e:
            print("[cleanup] error:", e)
        time.sleep(CLEANUP_EVERY_SECONDS)


threading.Thread(target=cleanup_old_metrics_loop, daemon=True).start()